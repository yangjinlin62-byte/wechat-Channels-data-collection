#!/usr/bin/env python3
"""Collect WeChat Channels account videos through WorldTreeTech and export XLSX."""

from __future__ import annotations

import argparse
import concurrent.futures
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any
from urllib import error, parse, request

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("Missing dependency: openpyxl is required to export .xlsx files.") from exc


BASE_URL = "https://www.worldtreetech.cn"
SEARCH_URL = f"{BASE_URL}/api/v2/wechat/video/search"
USER_INFO_URL = f"{BASE_URL}/api/v2/wechat/video/getUserInfo"
BALANCE_URL = f"{BASE_URL}/api/v2/user/balance"
DECRYPT_REPO_URL = "https://github.com/Evil0ctal/WeChat-Channels-Video-File-Decryption.git"
DEFAULT_TOOLS_DIR = Path.home() / ".codex" / "tools" / "wechat-channels-native-pipeline"
SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from native_asr import transcribe_audio_file

HEADERS = {
    "Content-Type": "application/json;charset=utf-8",
    "User-Agent": "Codex-WeChat-Video-Collector/1.0",
}

OUTPUT_HEADERS = [
    "达人昵称",
    "视频描述",
    "点赞量",
    "收藏量",
    "评论量",
    "分享量",
    "发布时间",
    "视频URL",
    "视频解密key",
]

FINAL_HEADERS = [
    "达人昵称",
    "视频描述",
    "视频文案",
    "点赞量",
    "收藏量",
    "评论量",
    "分享量",
    "发布时间",
]


class ApiError(RuntimeError):
    """Raised when WorldTreeTech returns a failed response."""


def get_api_key(args_key: str | None) -> str:
    key = (
        args_key
        or os.environ.get("WORLDTREE_API_KEY")
        or os.environ.get("WORLD_TREE_API_KEY")
        or read_windows_user_env("WORLDTREE_API_KEY")
        or read_windows_user_env("WORLD_TREE_API_KEY")
    )
    if not key:
        raise SystemExit(
            "API key missing. Set WORLDTREE_API_KEY or pass --key for this run."
        )
    return key


def read_windows_user_env(name: str) -> str | None:
    if os.name != "nt":
        return None
    try:
        import winreg

        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Environment") as key:
            value, _ = winreg.QueryValueEx(key, name)
            return str(value) if value else None
    except OSError:
        return None


def post_json(url: str, payload: dict[str, Any], timeout: int = 45) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = request.Request(url, data=body, headers=HEADERS, method="POST")
    try:
        with request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ApiError(f"HTTP {exc.code} from {url}: {detail}") from exc
    except error.URLError as exc:
        raise ApiError(f"Network error calling {url}: {exc}") from exc
    return parse_response(raw, url)


def get_json(url: str, params: dict[str, str], timeout: int = 30) -> dict[str, Any]:
    full_url = f"{url}?{parse.urlencode(params)}"
    req = request.Request(full_url, headers={"User-Agent": HEADERS["User-Agent"]}, method="GET")
    try:
        with request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ApiError(f"HTTP {exc.code} from {url}: {detail}") from exc
    except error.URLError as exc:
        raise ApiError(f"Network error calling {url}: {exc}") from exc
    return parse_response(raw, url)


def get_url(url: str, timeout: int = 5) -> bytes | None:
    req = request.Request(url, headers={"User-Agent": HEADERS["User-Agent"]}, method="GET")
    try:
        with request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except (error.HTTPError, error.URLError, TimeoutError):
        return None


def parse_response(raw: str, url: str) -> dict[str, Any]:
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ApiError(f"Invalid JSON from {url}: {raw[:500]}") from exc

    code = data.get("code")
    if code not in (None, 0, 200):
        msg = data.get("msg") or data.get("error") or "unknown error"
        raise ApiError(f"API error from {url}: code={code}, msg={msg}")
    return data


def search_accounts(key: str, account: str) -> list[dict[str, Any]]:
    result = post_json(SEARCH_URL, {"key": key, "name": account})
    accounts = result.get("data") or []
    if not isinstance(accounts, list):
        raise ApiError("Unexpected account search response: data is not a list.")
    return accounts


def choose_account(
    candidates: list[dict[str, Any]], account: str, candidate_index: int
) -> dict[str, Any]:
    if not candidates:
        raise SystemExit(f"No account candidates found for: {account}")

    exact = [
        item for item in candidates
        if str(item.get("nickname", "")).strip() == account.strip()
    ]
    if len(exact) == 1:
        return exact[0]

    index = candidate_index - 1
    if index < 0 or index >= len(candidates):
        raise SystemExit(f"--candidate-index must be between 1 and {len(candidates)}")
    return candidates[index]


def collect_videos(
    key: str,
    username: str,
    max_pages: int = 0,
    max_videos: int = 0,
    sleep_seconds: float = 0.3,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    videos: list[dict[str, Any]] = []
    account_info: dict[str, Any] = {}
    last_buffer = ""
    seen_buffers: set[str] = set()
    page = 0

    while True:
        page += 1
        payload = {"key": key, "username": username, "last_buffer": last_buffer}
        result = post_json(USER_INFO_URL, payload)
        data = result.get("data") or {}
        if not isinstance(data, dict):
            raise ApiError("Unexpected user info response: data is not an object.")

        if not account_info:
            account_info = data

        page_videos = data.get("video_list") or []
        if not isinstance(page_videos, list):
            raise ApiError("Unexpected user info response: video_list is not a list.")

        videos.extend(page_videos)
        if max_videos and len(videos) >= max_videos:
            videos = videos[:max_videos]
            break

        next_buffer = str(data.get("last_buffer") or "")
        continue_flag = data.get("continue_flag")
        if max_pages and page >= max_pages:
            break
        if continue_flag in (0, "0", False):
            break
        if not next_buffer or next_buffer in seen_buffers or next_buffer == last_buffer:
            break

        seen_buffers.add(next_buffer)
        last_buffer = next_buffer
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)

    return account_info, videos


def normalize_count(value: Any) -> int | str:
    if value is None or value == "":
        return ""
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def format_time(item: dict[str, Any]) -> str:
    create_time_str = item.get("create_time_str")
    if create_time_str:
        return str(create_time_str)

    raw = item.get("create_time")
    if raw in (None, ""):
        return ""
    try:
        timestamp = int(raw)
        if timestamp > 10_000_000_000:
            timestamp = timestamp // 1000
        return dt.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError, OSError):
        return str(raw)


def row_from_video(item: dict[str, Any], account_nickname: str) -> list[Any]:
    video_url = item.get("video_url") or item.get("url") or ""
    return [
        item.get("nickname") or account_nickname,
        item.get("title") or "",
        normalize_count(item.get("like_count", item.get("fav_count"))),
        normalize_count(item.get("fav_count")),
        normalize_count(item.get("comment_count")),
        normalize_count(item.get("forward_count")),
        format_time(item),
        video_url,
        item.get("decode_key") or "",
    ]


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip()
    cleaned = cleaned.rstrip(". ")
    return cleaned or "视频号账号"


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "视频号账号"
    return cleaned[:31]


def default_videos_dir(workbook_path: Path) -> Path:
    return workbook_path.with_suffix("").parent / f"{workbook_path.stem}-videos"


def default_decrypted_dir(workbook_path: Path) -> Path:
    return workbook_path.with_suffix("").parent / f"{workbook_path.stem}-decrypted"


def default_audio_dir(workbook_path: Path) -> Path:
    return workbook_path.with_suffix("").parent / f"{workbook_path.stem}-audio"


def default_transcripts_dir(workbook_path: Path) -> Path:
    return workbook_path.with_suffix("").parent / f"{workbook_path.stem}-transcripts"


def load_workbook_records(workbook_path: Path) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    records: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        record = {"_row": row_idx}
        for col_idx, header in enumerate(headers, start=1):
            if header:
                record[header] = ws.cell(row_idx, col_idx).value
        records.append(record)
    return records


def video_filename(record: dict[str, Any]) -> str:
    row_idx = int(record.get("_row") or 0) - 1
    title = str(record.get("视频描述") or record.get("达人昵称") or "video")
    name = safe_filename(title)[:60]
    return f"{row_idx:04d}-{name}.mp4"


def download_file(url: str, output_path: Path, timeout: int = 180) -> None:
    req = request.Request(url, headers={"User-Agent": HEADERS["User-Agent"]}, method="GET")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with request.urlopen(req, timeout=timeout) as resp, output_path.open("wb") as out:
        shutil.copyfileobj(resp, out)


def download_from_workbook(workbook_path: Path, videos_dir: Path | None = None) -> list[Path]:
    records = load_workbook_records(workbook_path)
    output_dir = videos_dir or default_videos_dir(workbook_path)
    downloaded: list[Path] = []
    for record in records:
        url = str(record.get("视频URL") or "").strip()
        if not url:
            print(f"Skip row {record.get('_row')}: missing 视频URL")
            continue
        output_path = output_dir / video_filename(record)
        if output_path.exists() and output_path.stat().st_size > 0:
            print(f"Exists: {output_path}")
        else:
            print(f"Downloading row {record.get('_row')}: {output_path.name}")
            download_file(url, output_path)
        downloaded.append(output_path)
    print(f"Downloaded/available videos: {len(downloaded)} in {output_dir}")
    return downloaded


def ensure_git_repo(repo_dir: Path) -> None:
    if (repo_dir / ".git").exists():
        return
    git = shutil.which("git")
    if not git:
        raise SystemExit("git is required to auto-install the WeChat video decryption tool.")
    repo_dir.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run([git, "clone", "--depth", "1", DECRYPT_REPO_URL, str(repo_dir)], check=True)


def run_checked(command: list[str], cwd: Path) -> None:
    print(f"Running: {' '.join(command)}")
    subprocess.run(command, cwd=str(cwd), check=True)


def which_command(*names: str) -> str | None:
    for name in names:
        path = shutil.which(name)
        if path:
            return path
    return None


def default_decrypt_tool_dir(workbook_path: Path | None = None) -> Path:
    return DEFAULT_TOOLS_DIR / "WeChat-Channels-Video-File-Decryption"


def ensure_decrypt_service(base_url: str, repo_dir: Path) -> None:
    if get_url(f"{base_url.rstrip('/')}/api/info", timeout=3):
        return

    ensure_git_repo(repo_dir)
    service_dir = repo_dir / "api-service"
    if not service_dir.exists():
        raise SystemExit(f"Decryption API service directory not found: {service_dir}")

    npm = which_command("npm.cmd", "npm.exe", "npm")
    node = which_command("node.exe", "node")
    if not npm or not node:
        raise SystemExit("Node.js and npm are required to auto-start the decryption service.")

    if not (service_dir / "node_modules").exists():
        run_checked([npm, "install"], service_dir)

    log_path = service_dir / "wechat-decrypt-api.log"
    log_file = log_path.open("ab")
    env = os.environ.copy()
    env.setdefault("PORT", parse.urlparse(base_url).port and str(parse.urlparse(base_url).port) or "3000")
    creationflags = subprocess.CREATE_NEW_PROCESS_GROUP if os.name == "nt" else 0
    subprocess.Popen(
        [node, "server.js"],
        cwd=str(service_dir),
        stdout=log_file,
        stderr=subprocess.STDOUT,
        stdin=subprocess.DEVNULL,
        env=env,
        creationflags=creationflags,
    )

    info_url = f"{base_url.rstrip('/')}/api/info"
    health_url = f"{base_url.rstrip('/')}/health"
    deadline = time.time() + 180
    while time.time() < deadline:
        if get_url(info_url, timeout=3):
            get_url(health_url, timeout=30)
            return
        time.sleep(2)
    raise SystemExit(f"Decryption service did not start in time. See log: {log_path}")


def decrypt_video_via_api(
    api_url: str,
    encrypted_path: Path,
    decode_key: str,
    output_path: Path,
    timeout: int = 600,
) -> None:
    boundary = f"----CodexWechatVideo{int(time.time() * 1000)}"
    video_bytes = encrypted_path.read_bytes()
    parts = [
        (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="decode_key"\r\n\r\n'
            f"{decode_key}\r\n"
        ).encode("utf-8"),
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="video"; filename="{encrypted_path.name}"\r\n'
            "Content-Type: video/mp4\r\n\r\n"
        ).encode("utf-8"),
        video_bytes,
        f"\r\n--{boundary}--\r\n".encode("utf-8"),
    ]
    body = b"".join(parts)
    req = request.Request(
        f"{api_url.rstrip('/')}/api/decrypt",
        data=body,
        method="POST",
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Content-Length": str(len(body)),
            "User-Agent": HEADERS["User-Agent"],
        },
    )
    try:
        with request.urlopen(req, timeout=timeout) as resp:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(resp.read())
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ApiError(f"Decrypt API failed for {encrypted_path.name}: HTTP {exc.code}: {detail}") from exc


def decrypt_from_workbook(
    workbook_path: Path,
    videos_dir: Path | None = None,
    decrypted_dir: Path | None = None,
    api_url: str = "http://localhost:3000",
    tool_dir: Path | None = None,
    download_missing: bool = False,
) -> list[Path]:
    if download_missing:
        download_from_workbook(workbook_path, videos_dir)

    records = load_workbook_records(workbook_path)
    source_dir = videos_dir or default_videos_dir(workbook_path)
    output_dir = decrypted_dir or default_decrypted_dir(workbook_path)
    ensure_decrypt_service(api_url, tool_dir or default_decrypt_tool_dir(workbook_path))

    decrypted: list[Path] = []
    for record in records:
        decode_key = str(record.get("视频解密key") or "").strip()
        if not decode_key:
            print(f"Skip row {record.get('_row')}: missing 视频解密key")
            continue
        encrypted_path = source_dir / video_filename(record)
        if not encrypted_path.exists():
            print(f"Skip row {record.get('_row')}: video file not found: {encrypted_path.name}")
            continue
        output_path = output_dir / encrypted_path.name
        if output_path.exists() and output_path.stat().st_size > 0:
            print(f"Exists: {output_path}")
        else:
            print(f"Decrypting row {record.get('_row')}: {encrypted_path.name}")
            decrypt_video_via_api(api_url, encrypted_path, decode_key, output_path)
        decrypted.append(output_path)
    print(f"Decrypted/available videos: {len(decrypted)} in {output_dir}")
    return decrypted


def resolve_ffmpeg() -> Path:
    env_value = os.environ.get("FFMPEG_BIN")
    candidates = []
    if env_value:
        candidates.append(Path(env_value))
    which_ffmpeg = shutil.which("ffmpeg") or shutil.which("ffmpeg.exe")
    if which_ffmpeg:
        candidates.append(Path(which_ffmpeg))
    candidates.extend(
        [
            DEFAULT_TOOLS_DIR / "ffmpeg" / "bin" / "ffmpeg.exe",
            Path(r"C:\ffmpeg\bin\ffmpeg.exe"),
            Path(r"C:\ProgramData\chocolatey\bin\ffmpeg.exe"),
        ]
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise SystemExit(
        "ffmpeg.exe not found. Run scripts\\install.ps1 or set FFMPEG_BIN to the full ffmpeg.exe path."
    )


def extract_audio(video_path: Path, audio_path: Path) -> None:
    audio_path.parent.mkdir(parents=True, exist_ok=True)
    ffmpeg = resolve_ffmpeg()
    command = [
        str(ffmpeg),
        "-hide_banner",
        "-loglevel",
        "error",
        "-y",
        "-i",
        str(video_path),
        "-vn",
        "-ac",
        "1",
        "-ar",
        "16000",
        "-b:a",
        "64k",
        str(audio_path),
    ]
    subprocess.run(command, check=True)


def transcribe_video(
    video_path: Path,
    transcript_path: Path,
    audio_path: Path,
    engine: str = "B",
    overwrite: bool = False,
) -> str:
    if transcript_path.exists() and transcript_path.stat().st_size > 0 and not overwrite:
        return transcript_path.read_text(encoding="utf-8").strip()

    if not audio_path.exists() or audio_path.stat().st_size == 0 or overwrite:
        extract_audio(video_path, audio_path)

    text, used_engine = transcribe_audio_file(audio_path, preferred_engine=engine)
    transcript_path.parent.mkdir(parents=True, exist_ok=True)
    transcript_path.write_text(text, encoding="utf-8")
    print(f"ASR engine {used_engine} succeeded for {video_path.name}.")
    return text.strip()


def ensure_workbook_column(ws: Any, header: str, width: int = 60) -> int:
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(1, col_idx).value or "").strip() == header:
            return col_idx

    col_idx = ws.max_column + 1
    ws.cell(1, col_idx).value = header
    ws.cell(1, col_idx).fill = PatternFill("solid", fgColor="1F4E78")
    ws.cell(1, col_idx).font = Font(color="FFFFFF", bold=True)
    ws.cell(1, col_idx).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = width
    return col_idx


def save_workbook_with_retry(wb: Any, workbook_path: Path, attempts: int = 8) -> None:
    for attempt in range(1, attempts + 1):
        try:
            wb.save(workbook_path)
            return
        except PermissionError:
            if attempt == attempts:
                raise
            time.sleep(1.5)


def transcribe_from_workbook(
    workbook_path: Path,
    video_dir: Path | None = None,
    transcript_dir: Path | None = None,
    audio_dir: Path | None = None,
    asrtools_dir: Path | None = None,
    engine: str = "B",
    asr_concurrency: int = 1,
    max_videos: int = 0,
    overwrite: bool = False,
) -> int:
    import openpyxl

    if asrtools_dir:
        print("--asrtools-dir is ignored by this native skill; using bundled ASR adapters and system ffmpeg.")
    source_dir = video_dir or default_decrypted_dir(workbook_path)
    output_dir = transcript_dir or default_transcripts_dir(workbook_path)
    work_audio_dir = audio_dir or default_audio_dir(workbook_path)

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    text_col = ensure_workbook_column(ws, "视频文案", width=80)
    records = load_workbook_records(workbook_path)

    jobs = []
    for record in records:
        row_idx = int(record.get("_row") or 0)
        if not overwrite and ws.cell(row_idx, text_col).value:
            continue

        video_path = source_dir / video_filename(record)
        if not video_path.exists():
            print(f"Skip row {row_idx}: decrypted video not found: {video_path.name}")
            continue

        transcript_path = output_dir / f"{video_path.stem}.txt"
        audio_path = work_audio_dir / f"{video_path.stem}.mp3"
        jobs.append((row_idx, video_path, transcript_path, audio_path))
        if max_videos and len(jobs) >= max_videos:
            break

    def run_job(job: tuple[int, Path, Path, Path]) -> tuple[int, str | None, str | None]:
        row_idx, video_path, transcript_path, audio_path = job
        print(f"Transcribing row {row_idx}: {video_path.name}")
        try:
            text = transcribe_video(
                video_path=video_path,
                transcript_path=transcript_path,
                audio_path=audio_path,
                engine=engine,
                overwrite=overwrite,
            )
        except (subprocess.CalledProcessError, RuntimeError) as exc:
            return row_idx, None, str(exc)
        return row_idx, text, None

    processed = 0
    failed: list[tuple[int, str]] = []
    workers = max(1, int(asr_concurrency or 1))
    if workers == 1:
        results = [run_job(job) for job in jobs]
    else:
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(run_job, job) for job in jobs]
            for future in concurrent.futures.as_completed(futures):
                results.append(future.result())

    for row_idx, text, err in sorted(results, key=lambda item: item[0]):
        if err:
            failed.append((row_idx, err))
            print(f"Skip row {row_idx}: ASR failed after fallbacks: {err}")
            continue
        ws.cell(row_idx, text_col).value = text
        ws.cell(row_idx, text_col).alignment = Alignment(wrap_text=True, vertical="top")
        processed += 1

    save_workbook_with_retry(wb, workbook_path)
    print(
        f"Transcribed and filled rows: {processed}; failed rows: {len(failed)}; "
        f"concurrency: {workers}; workbook: {workbook_path}"
    )
    return processed


def export_xlsx(
    videos: list[dict[str, Any]],
    output_path: Path,
    account_nickname: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(account_nickname)
    ws.append(OUTPUT_HEADERS)

    for item in videos:
        ws.append(row_from_video(item, account_nickname))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [18, 58, 12, 12, 12, 12, 20, 80, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[7].alignment = Alignment(wrap_text=False, vertical="top")
        url = row[7].value
        if url:
            row[7].hyperlink = str(url)
            row[7].style = "Hyperlink"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_final_xlsx(source_workbook: Path, output_path: Path, account_nickname: str) -> None:
    import openpyxl

    source_wb = openpyxl.load_workbook(source_workbook, data_only=True)
    source_ws = source_wb.active
    headers = [str(source_ws.cell(1, col).value or "").strip() for col in range(1, source_ws.max_column + 1)]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(account_nickname)
    ws.append(FINAL_HEADERS)

    for row_idx in range(2, source_ws.max_row + 1):
        row = []
        for header in FINAL_HEADERS:
            col_idx = header_index.get(header)
            row.append(source_ws.cell(row_idx, col_idx).value if col_idx else "")
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [18, 58, 80, 12, 12, 12, 12, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    save_workbook_with_retry(wb, output_path)


def cleanup_path(path: Path) -> None:
    if not path.exists():
        return
    if path.is_dir():
        shutil.rmtree(path)
    else:
        path.unlink()


def resolve_collection_target(args: argparse.Namespace, key: str) -> tuple[str, str, dict[str, Any]]:
    selected: dict[str, Any] = {}
    if args.username:
        return args.username, args.account or "视频号账号", selected

    if not args.account:
        raise SystemExit("Provide --account or --username.")

    candidates = search_accounts(key, args.account)
    if args.list_accounts:
        print_candidates(candidates)
        raise SystemExit(0)

    selected = choose_account(candidates, args.account, args.candidate_index)
    username = str(selected.get("username") or "")
    if not username:
        raise SystemExit("Selected account has no username.")
    return username, str(selected.get("nickname") or args.account), selected


def run_full_pipeline(args: argparse.Namespace) -> Path:
    key = get_api_key(args.key)
    username, requested_name, selected = resolve_collection_target(args, key)

    account_info, videos = collect_videos(
        key=key,
        username=username,
        max_pages=args.max_pages,
        max_videos=args.max_videos,
        sleep_seconds=args.sleep,
    )
    account_nickname = str(
        account_info.get("nickname") or selected.get("nickname") or requested_name
    )
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_name = f"{safe_filename(account_nickname)}-视频号账号数据.xlsx"
    final_path = output_dir / output_name
    work_root = (
        Path(args.pipeline_work_dir).expanduser().resolve()
        if args.pipeline_work_dir
        else output_dir / f"{safe_filename(account_nickname)}-视频号账号数据-work"
    )
    raw_path = work_root / f"{safe_filename(account_nickname)}-raw.xlsx"
    encrypted_dir = work_root / "encrypted-videos"
    audio_dir = work_root / "audio"
    transcript_dir = work_root / "transcripts"
    decrypted_dir = output_dir / f"{safe_filename(account_nickname)}-视频号账号数据-decrypted"

    print(f"Collected rows: {len(videos)}")
    export_xlsx(videos, raw_path, account_nickname)
    download_from_workbook(raw_path, encrypted_dir)
    decrypted = decrypt_from_workbook(
        raw_path,
        videos_dir=encrypted_dir,
        decrypted_dir=decrypted_dir,
        api_url=args.decrypt_api_url,
        tool_dir=Path(args.decrypt_tool_dir).expanduser().resolve() if args.decrypt_tool_dir else None,
    )
    transcribe_from_workbook(
        raw_path,
        video_dir=decrypted_dir,
        transcript_dir=transcript_dir,
        audio_dir=audio_dir,
        asrtools_dir=Path(args.asrtools_dir).expanduser().resolve() if args.asrtools_dir else None,
        engine=args.asr_engine,
        asr_concurrency=args.asr_concurrency,
        max_videos=args.max_transcribe,
        overwrite=args.overwrite_transcripts,
    )
    export_final_xlsx(raw_path, final_path, account_nickname)

    if not args.keep_encrypted_videos and len(decrypted) == len(videos):
        cleanup_path(encrypted_dir)
        print(f"Deleted encrypted video directory: {encrypted_dir}")
    elif len(decrypted) != len(videos):
        print("Encrypted videos kept because not every row was decrypted.")

    if not args.keep_audio_files:
        cleanup_path(audio_dir)
        print(f"Deleted audio directory: {audio_dir}")

    if not args.keep_pipeline_raw:
        cleanup_path(raw_path)
        print(f"Deleted raw intermediate workbook: {raw_path}")

    print(f"Final workbook: {final_path}")
    print(f"Decrypted videos kept at: {decrypted_dir}")
    return final_path


def print_candidates(candidates: list[dict[str, Any]]) -> None:
    if not candidates:
        print("No candidates found.")
        return
    for idx, item in enumerate(candidates, start=1):
        nickname = item.get("nickname", "")
        username = item.get("username", "")
        signature = item.get("signature", "")
        print(f"{idx}. {nickname} | {username} | {signature}")


def check_balance(key: str) -> None:
    result = get_json(BALANCE_URL, {"key": key})
    msg = result.get("msg") or result.get("message") or "success"
    balance = result.get("balance")
    print(f"Balance check: {msg}; balance={balance}")


def run_doctor(args: argparse.Namespace) -> int:
    failures = 0

    def report(name: str, ok: bool, detail: str) -> None:
        nonlocal failures
        status = "OK" if ok else "FAIL"
        print(f"[{status}] {name}: {detail}")
        if not ok:
            failures += 1

    report("Python", True, sys.executable)
    report("openpyxl", True, "imported")
    try:
        import requests as _requests  # noqa: F401

        report("requests", True, "imported")
    except ImportError as exc:
        report("requests", False, str(exc))

    try:
        ffmpeg = resolve_ffmpeg()
        report("ffmpeg", True, str(ffmpeg))
    except SystemExit as exc:
        report("ffmpeg", False, str(exc))

    if os.environ.get("WORLDTREE_API_KEY") or os.environ.get("WORLD_TREE_API_KEY") or args.key:
        report("WorldTreeTech key", True, "set")
    else:
        print("[WARN] WorldTreeTech key: missing; register at https://www.worldtreetech.cn/ before collection")

    git = which_command("git")
    node = which_command("node")
    npm = which_command("npm.cmd", "npm")
    report("git", bool(git), git or "missing")
    report("node", bool(node), node or "missing")
    report("npm", bool(npm), npm or "missing")

    tool_dir = Path(args.decrypt_tool_dir).expanduser().resolve() if args.decrypt_tool_dir else default_decrypt_tool_dir()
    report("decrypt tool", (tool_dir / "api-service" / "package.json").exists(), str(tool_dir / "api-service"))
    return 0 if failures == 0 else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Collect WeChat Channels account videos and export an .xlsx workbook."
    )
    parser.add_argument("--account", help="Account nickname or search keyword.")
    parser.add_argument("--username", help="WorldTreeTech/WeChat Channels username; skips account search.")
    parser.add_argument("--key", help="WorldTreeTech API key. Prefer WORLDTREE_API_KEY instead.")
    parser.add_argument("--output-dir", default=".", help="Directory for the output workbook.")
    parser.add_argument("--doctor", action="store_true", help="Check native pipeline dependencies and exit.")
    parser.add_argument("--full-pipeline", action="store_true", help="Collect, download, decrypt, transcribe, clean up, and export final 8-column workbook.")
    parser.add_argument("--pipeline-work-dir", help="Directory for full-pipeline intermediate files.")
    parser.add_argument("--candidate-index", type=int, default=1, help="1-based candidate index when search is ambiguous.")
    parser.add_argument("--list-accounts", action="store_true", help="Only list account search candidates.")
    parser.add_argument("--max-pages", type=int, default=0, help="Maximum pages to fetch; 0 means all available pages.")
    parser.add_argument("--max-videos", type=int, default=0, help="Maximum videos to export; 0 means no limit.")
    parser.add_argument("--sleep", type=float, default=0.3, help="Seconds to wait between paginated calls.")
    parser.add_argument("--check-balance", action="store_true", help="Check API balance and exit.")
    parser.add_argument("--download-from-xlsx", help="Download videos from an exported workbook's 视频URL column.")
    parser.add_argument("--videos-dir", help="Directory for encrypted downloaded videos.")
    parser.add_argument("--decrypt-from-xlsx", help="Decrypt downloaded videos using workbook 视频解密key values.")
    parser.add_argument("--decrypted-dir", help="Directory for decrypted videos.")
    parser.add_argument("--download-missing", action="store_true", help="When decrypting, download missing encrypted videos first.")
    parser.add_argument("--decrypt-api-url", default="http://localhost:3000", help="Local WeChat decrypt API base URL.")
    parser.add_argument("--decrypt-tool-dir", help="Directory for Evil0ctal's decrypt tool repo.")
    parser.add_argument("--transcribe-from-xlsx", help="Transcribe decrypted videos and fill the 视频文案 workbook column.")
    parser.add_argument("--transcribe-after-decrypt", action="store_true", help="After decryption, transcribe videos and fill 视频文案.")
    parser.add_argument("--transcribe-video-dir", help="Directory containing playable decrypted videos for transcription.")
    parser.add_argument("--transcripts-dir", help="Directory for transcript .txt files.")
    parser.add_argument("--audio-dir", help="Directory for extracted audio files.")
    parser.add_argument("--asrtools-dir", help="Deprecated compatibility option; ignored by the native skill.")
    parser.add_argument("--asr-engine", default="B", choices=["B", "J", "K"], help="Native ASR engine preference: B, J, or K. Failed items fall back automatically.")
    parser.add_argument("--asr-concurrency", type=int, default=2, help="ASR concurrency. Default 2 for stability.")
    parser.add_argument("--max-transcribe", type=int, default=0, help="Maximum videos to transcribe; 0 means no limit.")
    parser.add_argument("--overwrite-transcripts", action="store_true", help="Overwrite existing 视频文案 and transcript files.")
    parser.add_argument("--keep-encrypted-videos", action="store_true", help="Full pipeline: keep encrypted downloaded videos.")
    parser.add_argument("--keep-audio-files", action="store_true", help="Full pipeline: keep extracted audio files.")
    parser.add_argument("--keep-pipeline-raw", action="store_true", help="Full pipeline: keep raw intermediate workbook with video URL and decode_key.")
    return parser


def main(argv: list[str] | None = None) -> int:
    configure_stdout()
    args = build_parser().parse_args(argv)

    try:
        if args.doctor:
            return run_doctor(args)

        if args.full_pipeline:
            run_full_pipeline(args)
            return 0

        if args.download_from_xlsx:
            workbook_path = Path(args.download_from_xlsx).expanduser().resolve()
            videos_dir = Path(args.videos_dir).expanduser().resolve() if args.videos_dir else None
            download_from_workbook(workbook_path, videos_dir)
            return 0

        if args.decrypt_from_xlsx:
            workbook_path = Path(args.decrypt_from_xlsx).expanduser().resolve()
            videos_dir = Path(args.videos_dir).expanduser().resolve() if args.videos_dir else None
            decrypted_dir = Path(args.decrypted_dir).expanduser().resolve() if args.decrypted_dir else None
            tool_dir = Path(args.decrypt_tool_dir).expanduser().resolve() if args.decrypt_tool_dir else None
            decrypt_from_workbook(
                workbook_path=workbook_path,
                videos_dir=videos_dir,
                decrypted_dir=decrypted_dir,
                api_url=args.decrypt_api_url,
                tool_dir=tool_dir,
                download_missing=args.download_missing,
            )
            if args.transcribe_after_decrypt:
                transcript_video_dir = (
                    Path(args.transcribe_video_dir).expanduser().resolve()
                    if args.transcribe_video_dir
                    else decrypted_dir
                )
                transcribe_from_workbook(
                    workbook_path=workbook_path,
                    video_dir=transcript_video_dir,
                    transcript_dir=Path(args.transcripts_dir).expanduser().resolve() if args.transcripts_dir else None,
                    audio_dir=Path(args.audio_dir).expanduser().resolve() if args.audio_dir else None,
                    asrtools_dir=Path(args.asrtools_dir).expanduser().resolve() if args.asrtools_dir else None,
                    engine=args.asr_engine,
                    asr_concurrency=args.asr_concurrency,
                    max_videos=args.max_transcribe,
                    overwrite=args.overwrite_transcripts,
                )
            return 0

        if args.transcribe_from_xlsx:
            workbook_path = Path(args.transcribe_from_xlsx).expanduser().resolve()
            transcribe_from_workbook(
                workbook_path=workbook_path,
                video_dir=Path(args.transcribe_video_dir).expanduser().resolve() if args.transcribe_video_dir else None,
                transcript_dir=Path(args.transcripts_dir).expanduser().resolve() if args.transcripts_dir else None,
                audio_dir=Path(args.audio_dir).expanduser().resolve() if args.audio_dir else None,
                asrtools_dir=Path(args.asrtools_dir).expanduser().resolve() if args.asrtools_dir else None,
                engine=args.asr_engine,
                asr_concurrency=args.asr_concurrency,
                max_videos=args.max_transcribe,
                overwrite=args.overwrite_transcripts,
            )
            return 0

        key = get_api_key(args.key)

        if args.check_balance:
            check_balance(key)
            return 0

        if not args.account and not args.username:
            raise SystemExit("Provide --account or --username.")

        username, requested_name, selected = resolve_collection_target(args, key)

        account_info, videos = collect_videos(
            key=key,
            username=username,
            max_pages=args.max_pages,
            max_videos=args.max_videos,
            sleep_seconds=args.sleep,
        )
        account_nickname = str(
            account_info.get("nickname") or selected.get("nickname") or requested_name
        )
        output_name = f"{safe_filename(account_nickname)}-视频号账号数据.xlsx"
        output_path = Path(args.output_dir).expanduser().resolve() / output_name
        export_xlsx(videos, output_path, account_nickname)
        print(f"Exported {len(videos)} rows to {output_path}")
        return 0
    except ApiError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


def configure_stdout() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, ValueError):
            pass


if __name__ == "__main__":
    raise SystemExit(main())
